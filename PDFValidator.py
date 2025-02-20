import os
import fitz
import openpyxl
import datetime
import shutil
import cv2
import numpy as np
from PIL import Image



def normalize_text(text):
    return ' '.join(text.split())

def get_position_lable(x, y, page_width, page_height):
    vertical_position = ''
    horizontal_position = ''

    if y < page_height / 3:
        vertical_position = 'top'
    elif y < 2 * page_height / 3:
        vertical_position = 'middle'
    else:
        vertical_position = 'bottom'

    if x < page_width / 3:
        horizontal_position = 'left'
    elif x < 2 * page_width / 3:
        horizontal_position = 'center'
    else:
        horizontal_position = 'right'

    if horizontal_position == 'center':
        return vertical_position
    else:
        return f"{vertical_position} {horizontal_position}"

def image_to_numpy(image_path):
    image = Image.open(image_path).convert('RGB')
    return  np.array(image)

def search_image_in_pdf(page, image_array):
    pix = page.get_pixmap()
    page_image = Image.frombytes("RGB", [pix.width, pix.height], pix.sample)
    page_array = np.array(page_image)

    #Convert images to greay scale
    page_gray = cv2.cvtColor(page_array, cv2.COLOR_BGR2GRAY)
    template_gray = cv2.cvtColor(image_array, cv2.COLOR_BGR2GRAY)

    #Inizialize SIFT detector

    sift = cv2.SIFT_create()

    # Find the keypoints and descriptors with SIFT

    kp1, des1 = sift.detectAndCompute(template_gray, None)
    kp2, des2 = sift.detectAndCompute(page_gray, None)

    #Flann parameters

    FLANN_INDEX_KDTREE = 1
    index_params = dict(algorithm = FLANN_INDEX_KDTREE, trees = 5)
    search_params = dict(checks = 50)

    flann = cv2.FlannBasedMatcher(index_params, search_params)

    matches = flann.knnMatch(des1, des2, k=2)

    #store all the good matches as per lowe's ratio test.

    good_matches = []
    for m,n in matches:
        if m.distance < 0.7 * n.distance:
            good_matches.append(m)

    #check if matches are found
    if len(good_matches) > 10:
        #get co-ordinates of matched img
        src_pts =np.float32([kp1[m.queryIdx].pt for m in good_matches]).reshape(-1, 2)
        dst_pts =np.float32([kp2[m.queryIdx].pt for m in good_matches]).reshape(-1, 2)
        x, y = np.mean(dst_pts, axis=0)
        return True, (x, y)
    return False, None


def validate_pdf_strings(pdf_type, pdf_path, search_string, expected_position, sheet):
    try:
        document = fitz.open(pdf_path)
        pdf_name = os.path.basename(pdf_path)
        normalized_search_string = normalize_text(search_string)
        for page_num in range(len(document)):
            page = document.load_page(page_num)
            text_instances = page.search_for(normalized_search_string)
            page_height = page.rect.height
            page_width = page.rect.width
            reported_positions = set()
            if text_instances:
                for inst in text_instances:
                    position_label = get_position_lable(inst.x0, inst.y0, page_width, page_height)
                    position_key = (page_num, round(inst.x0, 2), round(inst.y0,2)) #use rounded co-ordinates for precesion
                    if position_key not in reported_positions:
                        reported_positions.add(position_key)
                        coordinates = f"({inst.x0}, {inst.y0})"
                        status = 'Pass' if position_label == expected_position else 'Fail'
                        sheet.append([pdf_type, pdf_path, pdf_name, page_num +1, 'Found', coordinates, position_label, status])
                        print(f"Processing {pdf_type}, page {page_num + 1}: Found at {coordinates} {position_label} - {status}")
                    else:
                        sheet.append([pdf_type, pdf_path, pdf_name, page_num + 1, 'Not Found', 'Not Found', 'Not Found'])
                        print(f"Processing {pdf_type}, page {page_num + 1}: Not Found at ")

    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")


def validate_pdf_images(pdf_type, pdf_path, image_path,expected_position, sheet):
    try:
        document = fitz.open(pdf_path)
        pdf_name = os.path.basename(pdf_path)
        image_array = image_to_numpy(image_path)
        for page_num in range(len(document)):
            page = document.load_page(page_num)
            found, coordinates = search_image_in_pdf(page, image_array)
            if found:
                x, y = coordinates
                page_height = page.rect.height
                page_width = page.rect.width
                position_label = get_position_lable(x, y, page_width, page_height)
                status = 'Pass' if position_label == expected_position else 'Fail'
                sheet.append([pdf_type, pdf_path, pdf_name, page_num + 1, 'Found', f"({x}, {y})", position_label, status])
                print(f"Processing {pdf_type}, page {page_num + 1}: Image Found at ({x}, {y}) {position_label} - {status}")

            else:
                sheet.append([pdf_type, pdf_path, pdf_name, page_num + 1, 'Not Found', 'Image Not Found', 'Image Not Found', 'Fail'])
                print(f"Processing {pdf_type}, page {page_num + 1}: Image Not Found - Fail ")

    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")

def validate_pdfs_from_runner(runner_file, output_folder):
    #Move existing report files to the Archive Reports folder
    archive_folder = os.path.join(output_folder,'Archive Reports')
    os.makedirs(archive_folder, exist_ok=True)
    for filename in os.listdir(output_folder):
        if filename.endswith('.xlsx'):
            try:
                shutil.move(os.path.join(output_folder, filename),os.path.join(archive_folder,filename))
            except PermissionError:
                print("Please close all the reports and re-run again.")
                return

    #Create a new report file with the current date and timestamp
    timestamp = datetime.datetime.now().strftime('%y%m%d_%H%M%S')
    output_excel = os.path.join(output_folder, f'validation_report{timestamp}.xlsx')

    workbook = openpyxl.Workbook()

    #Load the runner Excel file

    runner_wb = openpyxl.load_workbook(runner_file)
    runner_sheet = runner_wb.active

    sheets = {}
    sheet_counts = {}

    for row in runner_sheet.iter_rows(min_row=2, values_only=True):
        pdf_type, pdf_folder_path, run,data_type, expected_value, expected_position = row

        if run.lower() == 'yes':
            pdf_type_str = str(pdf_type) #convert pdf_type to string
            if pdf_type_str not in sheet_counts:
                sheet_counts[pdf_type_str] = 0
            sheet_counts[pdf_type_str] += 1
            sheet_name = f"{pdf_type_str}_{sheet_counts[pdf_type_str]}"
            sheets[sheet_name] = workbook.create_sheet(title=sheet_name)
            sheets[sheet_name].append(["PDF Type", "PDF Path", "PDF name", "Page Number", "Found Status", "Co-Ordinates", "Position", "Status"])
            for filename in os.listdir(pdf_folder_path):
                pdf_path = os.path.join(pdf_folder_path, filename)
                print(f"Processing {pdf_path}")
                if data_type.lower() == 'img':
                    validate_pdf_images(pdf_type,pdf_path,expected_value,expected_position,sheets[sheet_name])
                else:
                    validate_pdf_strings(pdf_type,pdf_path,expected_value,expected_position,sheets[sheet_name])

    #Remove the default sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    try:
        workbook.save(output_excel)
    except PermissionError:
        print("Please close all the reports and re-run again.")

if __name__ == "__main__":
    runner_file = "Input_Data/RunnerData.xlsx"
    output_folder = "Output_Result"

    validate_pdfs_from_runner(runner_file, output_folder)







